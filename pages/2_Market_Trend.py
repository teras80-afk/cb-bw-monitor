"""
페이지 2 — 시장 전체 CB/BW 신규 발행 동향
"""
import streamlit as st
import pandas as pd
import altair as alt
from datetime import datetime
from lib import require_dart_or_stop, fetch_market_cb_bw_recent

st.set_page_config(page_title="시장 동향", page_icon="📈", layout="wide")
st.title("📈 시장 CB/BW 발행 동향")
st.caption("최근 시장 전체 CB/BW 신규 발행 공시 (어떤 회사가 자금 조달을 시작했나)")

require_dart_or_stop()

st.markdown("---")

col_d, col_btn = st.columns([3, 1])
with col_d:
    days_back = st.selectbox("조회 기간",
                              [7, 14, 30, 60, 90],
                              index=2,
                              format_func=lambda d: f"최근 {d}일")
with col_btn:
    st.write("")  # spacer
    refresh = st.button("🔄 새로고침", use_container_width=True)

if refresh:
    st.cache_data.clear()
    st.rerun()

with st.spinner(f"최근 {days_back}일 시장 전체 CB/BW 공시 조회 중..."):
    df = fetch_market_cb_bw_recent(days_back)

if df.empty:
    st.info("해당 기간 내 CB/BW 발행 공시 없음 (또는 DART API 응답 오류)")
    st.stop()

# 핵심 지표
n_total = len(df)
n_cb = (df["사채종류"] == "CB").sum() if "사채종류" in df.columns else 0
n_bw = (df["사채종류"] == "BW").sum() if "사채종류" in df.columns else 0
n_companies = (df["corp_name"].nunique()
               if "corp_name" in df.columns else 0)

m1, m2, m3, m4 = st.columns(4)
with m1:
    st.metric("총 공시 건수", f"{n_total}건")
with m2:
    st.metric("CB 발행", f"{n_cb}건")
with m3:
    st.metric("BW 발행", f"{n_bw}건")
with m4:
    st.metric("발행 회사 수", f"{n_companies}개사")

st.markdown("---")

# 일자별 추이 차트
if "rcept_dt" in df.columns:
    df_chart = df.copy()
    df_chart["접수일"] = pd.to_datetime(df_chart["rcept_dt"].astype(str),
                                       format="%Y%m%d", errors="coerce")
    df_chart = df_chart.dropna(subset=["접수일"])
    if len(df_chart) > 0:
        st.markdown("#### 📊 일자별 발행 공시 추이")
        daily = (df_chart.groupby([df_chart["접수일"].dt.date, "사채종류"])
                 .size().reset_index(name="건수"))
        daily.columns = ["접수일", "사채종류", "건수"]
        daily["접수일"] = pd.to_datetime(daily["접수일"])

        chart = alt.Chart(daily).mark_bar().encode(
            x=alt.X("접수일:T", title=""),
            y=alt.Y("건수:Q", title="공시 건수"),
            color=alt.Color("사채종류:N",
                            scale=alt.Scale(domain=["CB", "BW"],
                                            range=["#3498db", "#e67e22"])),
            tooltip=["접수일:T", "사채종류:N", "건수:Q"],
        ).properties(height=280)
        st.altair_chart(chart, use_container_width=True)

st.markdown("---")

# 공시 리스트
st.markdown("#### 📋 발행 공시 리스트")

# 시장 필터
filter_kind = st.radio("사채종류 필터", ["전체", "CB만", "BW만"],
                       horizontal=True, index=0)
df_show = df.copy()
if filter_kind == "CB만":
    df_show = df_show[df_show["사채종류"] == "CB"]
elif filter_kind == "BW만":
    df_show = df_show[df_show["사채종류"] == "BW"]

if df_show.empty:
    st.info("해당 조건에 맞는 공시 없음")
else:
    show_cols = []
    for c in ["rcept_dt", "corp_name", "stock_code", "사채종류", "report_nm"]:
        if c in df_show.columns:
            show_cols.append(c)
    rename_map = {
        "rcept_dt": "접수일",
        "corp_name": "회사명",
        "stock_code": "종목코드",
        "report_nm": "공시명",
    }
    df_view = (df_show[show_cols].rename(columns=rename_map)
               .sort_values("접수일", ascending=False)
               .reset_index(drop=True))
    st.caption(f"표시: **{len(df_view)}건**")
    st.dataframe(df_view, use_container_width=True, hide_index=True)

    # ─── 엑셀 다운로드 (종합 리포트) ───
    st.markdown("#### 📥 엑셀 다운로드")
    st.caption("위 데이터를 엑셀 파일로 받아 정기 보관·분석에 활용하세요. "
               "파일 안에 DART 원문 URL이 클릭 가능한 링크로 포함됩니다.")

    @st.cache_data(ttl=600)
    def build_excel_report(df_input: pd.DataFrame) -> bytes:
        """엑셀 리포트 생성 — 회사명+코드+종류+공시일+공시명+DART URL 포함"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "CB_BW 발행 리포트"

        # 헤더
        headers = ["접수일", "회사명", "종목코드", "사채종류", "공시명", "DART 원문 URL"]
        thin = Side(border_style="thin", color="BFBFBF")
        border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all

        # 데이터 행
        body_font = Font(name="맑은 고딕", size=10)
        link_font = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")

        for row_idx, (_, row) in enumerate(df_input.iterrows(), start=2):
            rcept_dt = str(row.get("rcept_dt", "")).strip()
            # YYYYMMDD → YYYY-MM-DD
            if len(rcept_dt) == 8 and rcept_dt.isdigit():
                rcept_dt_fmt = f"{rcept_dt[:4]}-{rcept_dt[4:6]}-{rcept_dt[6:8]}"
            else:
                rcept_dt_fmt = rcept_dt

            corp_name = str(row.get("corp_name", "—"))
            stock_code = str(row.get("stock_code", "—"))
            kind = str(row.get("사채종류", "—"))
            report_nm = str(row.get("report_nm", "—"))
            url = str(row.get("원문URL", ""))

            values = [rcept_dt_fmt, corp_name, stock_code, kind, report_nm, url]
            for col_idx, v in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                cell.font = body_font
                cell.border = border_all
                cell.alignment = Alignment(vertical="center", wrap_text=False)

            # URL을 클릭 가능한 하이퍼링크로
            if url and url != "#":
                link_cell = ws.cell(row=row_idx, column=6)
                link_cell.hyperlink = url
                link_cell.font = link_font

        # 컬럼 너비 자동 조정
        widths = [12, 24, 10, 10, 50, 60]
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # 첫 행 고정
        ws.freeze_panes = "A2"

        # 메타 정보 시트
        meta_ws = wb.create_sheet(title="리포트 정보")
        today = datetime.now().strftime("%Y-%m-%d %H:%M")
        meta = [
            ("리포트 제목", "CB/BW 발행 종합 리포트"),
            ("생성일시", today),
            ("총 건수", f"{len(df_input)}건"),
            ("CB 건수", f"{(df_input['사채종류'] == 'CB').sum()}건" if "사채종류" in df_input.columns else "-"),
            ("BW 건수", f"{(df_input['사채종류'] == 'BW').sum()}건" if "사채종류" in df_input.columns else "-"),
            ("발행 회사 수", f"{df_input['corp_name'].nunique()}개사" if "corp_name" in df_input.columns else "-"),
            ("데이터 출처", "DART 전자공시 (주요사항보고)"),
            ("주의", "정정공시·취소공시는 별도 표시되지 않으므로 DART 원문 교차 확인 필요"),
        ]
        for row_idx, (k, v) in enumerate(meta, 1):
            ws_k = meta_ws.cell(row=row_idx, column=1, value=k)
            ws_v = meta_ws.cell(row=row_idx, column=2, value=v)
            ws_k.font = Font(name="맑은 고딕", size=10, bold=True)
            ws_v.font = Font(name="맑은 고딕", size=10)
        meta_ws.column_dimensions["A"].width = 18
        meta_ws.column_dimensions["B"].width = 60

        # 바이트로 변환
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    try:
        excel_bytes = build_excel_report(df_show)
        filename = f"CB_BW_발행리포트_{datetime.now():%Y%m%d}.xlsx"
        st.download_button(
            label="📊 엑셀 파일 다운로드",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=False,
        )
        st.caption(f"💡 파일명: {filename} · 시트 2개 (발행 리포트 + 리포트 정보)")
    except Exception as e:
        st.error(f"엑셀 생성 실패: {e}")

    # DART 원문 링크
    with st.expander("🔗 DART 원문 바로가기"):
        for _, row in df_show.iterrows():
            rd = row.get("rcept_dt", "—")
            cn = row.get("corp_name", "—")
            rn = row.get("report_nm", "—")
            url = row.get("원문URL", "#")
            st.markdown(f"- [{rd}] **{cn}** — {rn} → [DART 원문]({url})")

st.markdown("---")
st.caption(
    "📌 DART 주요사항보고(kind='B') 중 'CB/BW 발행결정'만 필터링한 결과입니다. "
    "정정공시·취소공시는 별도 표시되지 않으니 DART 원문에서 최종 확인하세요."
)
